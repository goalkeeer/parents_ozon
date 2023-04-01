import csv
import traceback
from collections import defaultdict
from datetime import datetime
from decimal import Decimal
from itertools import islice
from pathlib import Path

import openpyxl
import xlsxwriter
from tqdm import tqdm

from .browsers import BrowserException
from program import settings
from .services import TradeLockService
from .utils import write_log, write_interim_result


def get_item_from_products_csv() -> dict:
    items = defaultdict(dict)
    with open('products.csv', encoding="utf8") as file:
        csv_reader = csv.reader(file, delimiter=';')
        for row in islice(csv_reader, 1, None):
            name = row[0]
            items[name] = {
                'stock_count': row[20],
                'reserve_count': row[21],
                'stock_amount': row[22]
            }
    return items


def get_items_from_excel():
    workbook = openpyxl.load_workbook('products.csv')
    worksheet = workbook.worksheets[1]

    return [worksheet.cell(r, 1).value for r in range(5, worksheet.max_row+1)]


def warning_item(values):
    stock_count = int(values['stock_count']) + int(values['reserve_count'])
    if stock_count + 5 > int(values['count']):
        return True

    if Decimal(values['stock_amount']) * Decimal(1 - settings.AMOUNT_DIFFERENCE) < Decimal(values['price']):
        return True

    return False


def write_excel_result(items):
    Path(settings.RESULT_PATH).mkdir(parents=True, exist_ok=True)
    time = datetime.now().strftime('%d-%mT%H:%M')
    file_path = Path(settings.RESULT_PATH, f'result_{time}.xlsx')
    headers = ('Наименование', 'Кол-во на озоне', 'Зарезервировано на озоне',
               'Кол-во на сайте', 'Цена на Озоне', 'Цена на сайте')
    workbook = xlsxwriter.Workbook(file_path, {'constant_memory': True})
    bold = workbook.add_format({'bold': True})

    bg_red = workbook.add_format()
    bg_red.set_pattern(1)
    bg_red.set_bg_color('red')

    bg_yellow = workbook.add_format()
    bg_yellow.set_pattern(1)
    bg_yellow.set_bg_color('yellow')

    worksheet = workbook.add_worksheet()

    for i, header in enumerate(headers):
        worksheet.write(0, i, header, bold)

    for i, (name, values) in enumerate(items.items()):
        bg = None
        if 'error' in values:
            row = (name, values['error'])
            bg = bg_red
        else:
            row = (name, values['stock_count'], values['reserve_count'],
                   values['count'], values['stock_amount'], values['price'])
            try:
                is_warning = warning_item(values)
            except Exception as error:
                print(error)
                tb = traceback.format_exc()
                write_log(str(tb))
                is_warning = True

            if is_warning:
                bg = bg_yellow
        for n, value in enumerate(row):
            worksheet.write(i + 1, n, value, bg)

    workbook.close()


def run_program():
    items = get_item_from_products_csv()

    auth = (settings.TRADE_LOCK_USERNAME, settings.TRADE_LOCK_PASSWORD)
    service = TradeLockService(*auth)
    for item_name in tqdm(items.keys()):
        try:
            data = service.get_item_data(item_name)
        except BrowserException as error:
            data = {'error': str(error)}
        except Exception as error:
            print(error)
            tb = traceback.format_exc()
            write_log(str(tb))
            data = {'error': 'непонятная ошибка'}

        items[item_name].update(data)
        write_interim_result(f'{item_name}: {items[item_name]}')
    service.quit()

    write_excel_result(items)
